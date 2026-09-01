"""Convert STICS alternating-line text files to template-based XML files.

Plant/cultivar and soil files are supported.  A template is required
because the text format contains values but none of the XML hierarchy or
parameter metadata (formalism names, bounds, formats, and option choices).
"""

from __future__ import annotations

import argparse
import colorsys
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
import sqlite3
from typing import Dict, Iterable, List, Optional, Tuple, Union
import xml.etree.ElementTree as ET

from .stics_cultivar_file import SticsCultivarFile


PathLike = Union[str, Path]


@dataclass(frozen=True)
class TxtToXmlResult:
    """Summary of one STICS plant TXT-to-XML conversion."""

    destination: Path
    cultivars: int
    species_parameters_updated: int
    cultivar_parameters_updated: int
    template_species_defaults: Tuple[str, ...]
    template_cultivar_defaults: Tuple[str, ...]


@dataclass(frozen=True)
class SoilTxtToXmlResult:
    """Summary of a combined STICS soil TXT-to-XML conversion."""

    destination: Path
    soils: int
    soil_parameters_updated: int
    layer_parameters_updated: int


@dataclass(frozen=True)
class InitializationTxtToXmlResult:
    """Summary of one STICS initialization TXT-to-XML conversion."""

    destination: Path
    plants: int
    plant_values_updated: int
    soil_values_updated: int
    snow_values_updated: int


@dataclass(frozen=True)
class TechnicalTxtToXmlResult:
    """Summary of one STICS technical-management TXT-to-XML conversion."""

    destination: Path
    parameters_updated: int
    intervention_tables: int
    interventions: int


@dataclass(frozen=True)
class StationTxtToXmlResult:
    """Summary of one STICS station TXT-to-XML conversion."""

    destination: Path
    parameters_updated: int
    template_defaults: Tuple[str, ...]


@dataclass(frozen=True)
class UsmDirectoryToXmlResult:
    """Summary of a complete directory-of-USMs conversion."""

    destination: Path
    usms: int
    soils: int
    climate_files: int
    associated_xml_files: int
    observation_files: int = 0
    summary_workbook: Optional[Path] = None


@dataclass(frozen=True)
class GeneralParametersTxtToXmlResult:
    """Summary of one STICS general-parameters TXT-to-XML conversion."""

    destination: Path
    values_updated: int
    distinct_parameter_names: int


@dataclass(frozen=True)
class ObservationsResult:
    """Summary of STICS observation files generated from MasterInput."""

    destination: Path
    files: int
    rows: int


@dataclass(frozen=True)
class _SoilTxtData:
    name: str
    parameters: Dict[str, str]
    layers: Tuple[Dict[str, str], ...]


_SOIL_LINE_1 = (
    "numsol", "typsol", "argi", "norg", "profhum", "calc", "pH",
    "concseuil", "albedo", "q0", "ruisolnu", "obstarac", "pluiebat",
    "mulchbat", "zesx", "cfes", "z0solnu", "CsurNsol", "finert",
    "penterui",
)
_SOIL_LINE_2 = (
    "numsol", "codecailloux", "codemacropor", "codefente", "codrainage",
    "coderemontcap", "codenitrif", "codedenit",
)
_SOIL_LINE_3 = (
    "numsol", "profimper", "ecartdrain", "ksol", "profdrain", "capiljour",
    "humcapil", "profdenit", "vpotdenit",
)
_SOIL_LAYER = (
    "numsol", "epc", "HCCF", "HMINF", "DAF", "cailloux",
    "typecailloux", "infil", "epd",
)


def _parse_template(path: Path) -> ET.ElementTree:
    if not path.is_file():
        raise FileNotFoundError(path)
    if path.stat().st_size == 0:
        raise ValueError(f"{path}: XML template is empty")

    try:
        parser = ET.XMLParser(target=ET.TreeBuilder(insert_comments=True))
        return ET.parse(path, parser=parser)
    except ET.ParseError as exc:
        raise ValueError(f"{path}: invalid XML template: {exc}") from exc


def _parameter_map(elements: Iterable[ET.Element], section: str) -> Dict[str, ET.Element]:
    result: Dict[str, ET.Element] = {}
    for element in elements:
        name: Optional[str] = None
        if element.tag == "param":
            name = element.get("nom")
        elif element.tag == "option":
            name = element.get("nomParam")
        if not name:
            continue
        if name in result:
            raise ValueError(f"duplicate XML parameter {name!r} in {section}")
        result[name] = element
    return result


def _species_parameter_map(root: ET.Element) -> Dict[str, ET.Element]:
    cultivar_descendants = {
        id(element)
        for cultivar in root.iter("variete")
        for element in cultivar.iter()
    }
    return _parameter_map(
        (element for element in root.iter() if id(element) not in cultivar_descendants),
        "species template",
    )


def _cultivar_parameter_map(cultivar: ET.Element) -> Dict[str, ET.Element]:
    # Cultivar choices are encoded by the ``code`` attributes of their
    # parameters; unlike species options, they have no nomParam value in TXT.
    return _parameter_map(cultivar.iter("param"), f"cultivar {cultivar.get('nom')!r}")


def _set_xml_value(element: ET.Element, value: str) -> None:
    if element.tag == "option":
        element.set("choix", str(value))
    else:
        element.text = str(value)


def _check_parameter_sets(
    source_names: Iterable[str],
    template_names: Iterable[str],
    section: str,
    strict: bool,
) -> Tuple[str, ...]:
    source = set(source_names)
    template = set(template_names)
    unsupported = sorted(source - template)
    if unsupported:
        raise ValueError(
            f"{section}: parameters from TXT absent from XML template: "
            + ", ".join(unsupported)
        )

    defaults = tuple(sorted(template - source))
    if strict and defaults:
        raise ValueError(
            f"{section}: parameters from XML template absent from TXT: "
            + ", ".join(defaults)
        )
    return defaults


def _indent(element: ET.Element, level: int = 0) -> None:
    """Indent an ElementTree without requiring Python 3.9's ET.indent."""

    whitespace = "\n" + level * "  "
    child_whitespace = "\n" + (level + 1) * "  "
    children = list(element)
    if children:
        if not element.text or not element.text.strip():
            element.text = child_whitespace
        for child in children:
            _indent(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = child_whitespace
        children[-1].tail = whitespace


def _write_xml(
    root: ET.Element, destination: Path, *, standalone: Optional[bool] = None
) -> None:
    _indent(root)
    destination.parent.mkdir(parents=True, exist_ok=True)
    xml_body = ET.tostring(root, encoding="unicode", short_empty_elements=True)
    xml_body = xml_body.replace(" />", "/>")
    declaration = '<?xml version="1.0" encoding="UTF-8"'
    if standalone is not None:
        declaration += f' standalone="{"yes" if standalone else "no"}"'
    declaration += "?>\n"
    destination.write_text(declaration + xml_body, encoding="utf-8")
    ET.parse(destination)


def convert_plant_txt_to_xml(
    source_txt: PathLike,
    template_xml: PathLike,
    destination_xml: PathLike,
    *,
    template_cultivar: Optional[str] = None,
    strict: bool = True,
) -> TxtToXmlResult:
    """Convert one STICS plant/cultivar TXT file using an XML template.

    Species ``param`` text and ``option@choix`` values are replaced from the
    TXT species section.  Existing template varieties are removed, one
    structural variety block is cloned for every TXT ``codevar``, and its
    parameter values are replaced from that cultivar's TXT block.

    Parameters absent from the template always raise an error because their
    XML location is unknown.  With ``strict=True`` (the default), parameters
    present only in the template also raise; with ``strict=False`` their
    template defaults are retained and reported in the result.
    """

    source_path = Path(source_txt)
    template_path = Path(template_xml)
    destination_path = Path(destination_xml)

    source = SticsCultivarFile.read(source_path)
    tree = _parse_template(template_path)
    root = tree.getroot()
    if root.tag != "fichierplt":
        raise ValueError(
            f"{template_path}: expected a fichierplt root, found {root.tag!r}"
        )

    species_xml = _species_parameter_map(root)
    species_defaults = _check_parameter_sets(
        source.species_parameters,
        species_xml,
        "species",
        strict,
    )
    for name, value in source.species_parameters.items():
        _set_xml_value(species_xml[name], value)

    tables = list(root.iter("tv"))
    if len(tables) != 1:
        raise ValueError(
            f"{template_path}: expected exactly one tv cultivar table, found {len(tables)}"
        )
    table = tables[0]
    template_varieties = [child for child in list(table) if child.tag == "variete"]
    if not template_varieties:
        raise ValueError(f"{template_path}: no variete block in cultivar table")

    if template_cultivar is None:
        cultivar_template = template_varieties[0]
    else:
        matches = [
            cultivar
            for cultivar in template_varieties
            if cultivar.get("nom") == template_cultivar
        ]
        if not matches:
            raise ValueError(
                f"{template_path}: template cultivar {template_cultivar!r} not found"
            )
        cultivar_template = matches[0]

    template_cultivar_parameters = _cultivar_parameter_map(cultivar_template)
    cultivar_defaults: Optional[Tuple[str, ...]] = None
    for cultivar_name, parameters in source.cultivars.items():
        defaults = _check_parameter_sets(
            parameters,
            template_cultivar_parameters,
            f"cultivar {cultivar_name!r}",
            strict,
        )
        if cultivar_defaults is None:
            cultivar_defaults = defaults
        elif cultivar_defaults != defaults:
            raise ValueError("TXT cultivars do not contain the same parameter set")

    for cultivar in template_varieties:
        table.remove(cultivar)

    cultivar_parameter_updates = 0
    for cultivar_name, parameters in source.cultivars.items():
        cultivar = deepcopy(cultivar_template)
        cultivar.set("nom", cultivar_name)
        cultivar_xml = _cultivar_parameter_map(cultivar)
        for name, value in parameters.items():
            _set_xml_value(cultivar_xml[name], value)
            cultivar_parameter_updates += 1
        table.append(cultivar)

    table.set("nb_varietes", str(len(source.cultivars)))
    _write_xml(root, destination_path, standalone=False)
    return TxtToXmlResult(
        destination=destination_path,
        cultivars=len(source.cultivars),
        species_parameters_updated=len(source.species_parameters),
        cultivar_parameters_updated=cultivar_parameter_updates,
        template_species_defaults=species_defaults,
        template_cultivar_defaults=cultivar_defaults or (),
    )


def _read_soil_txt(path: Path, name: Optional[str] = None) -> _SoilTxtData:
    if not path.is_file():
        raise FileNotFoundError(path)
    lines = [line.split() for line in path.read_text(encoding="utf-8-sig").splitlines()
             if line.strip()]
    if len(lines) != 8:
        raise ValueError(f"{path}: expected 8 non-empty lines, found {len(lines)}")

    schemas = (_SOIL_LINE_1, _SOIL_LINE_2, _SOIL_LINE_3) + (_SOIL_LAYER,) * 5
    for number, (values, schema) in enumerate(zip(lines, schemas), start=1):
        if len(values) != len(schema):
            raise ValueError(
                f"{path}: line {number} contains {len(values)} values; "
                f"expected {len(schema)}"
            )
    numbers = {line[0] for line in lines}
    if len(numbers) != 1:
        raise ValueError(f"{path}: inconsistent numsol values: {sorted(numbers)}")

    parameters: Dict[str, str] = {}
    for values, schema in zip(lines[:3], schemas[:3]):
        parameters.update(
            (name, value) for name, value in zip(schema, values)
            if name not in {"numsol", "typsol"}
        )
    layers = tuple(
        {name: value for name, value in zip(_SOIL_LAYER, values) if name != "numsol"}
        for values in lines[3:]
    )
    return _SoilTxtData(name or path.stem, parameters, layers)


def convert_soil_txt_to_xml(
    source_files: Iterable[PathLike],
    template_xml: PathLike,
    destination_xml: PathLike,
    *,
    template_soil: Optional[str] = None,
    soil_names: Optional[Iterable[str]] = None,
) -> SoilTxtToXmlResult:
    """Combine one or more legacy STICS ``.sol`` files into one XML file.

    The positional correspondence follows the STICS ``Sol.f90`` reader.  The
    XML soil name is the source filename stem.  Each legacy file must contain
    the three soil-level lines followed by exactly five layer lines.
    """

    paths = [Path(path) for path in source_files]
    if not paths:
        raise ValueError("at least one soil TXT file is required")
    names = list(soil_names) if soil_names is not None else [path.stem for path in paths]
    if len(names) != len(paths):
        raise ValueError("soil_names must contain exactly one name per source file")
    if len(names) != len(set(names)):
        raise ValueError("soil source filenames must have unique stems")
    sources = [_read_soil_txt(path, name) for path, name in zip(paths, names)]

    template_path = Path(template_xml)
    destination_path = Path(destination_xml)
    tree = _parse_template(template_path)
    root = tree.getroot()
    if root.tag != "sols":
        raise ValueError(f"{template_path}: expected a sols root, found {root.tag!r}")
    template_soils = [child for child in list(root) if child.tag == "sol"]
    if not template_soils:
        raise ValueError(f"{template_path}: no sol block in template")
    if template_soil is None:
        soil_template = template_soils[0]
    else:
        matches = [soil for soil in template_soils if soil.get("nom") == template_soil]
        if not matches:
            raise ValueError(f"{template_path}: template soil {template_soil!r} not found")
        soil_template = matches[0]

    for soil in template_soils:
        root.remove(soil)
    # Some official templates keep disabled example soils as direct XML
    # comments (for example ``<!--sol nom="solcarotte"> ...``).  They are
    # template data too and must not leak into the generated soil database.
    for child in list(root):
        # ElementTree exposes the content without the surrounding ``<!--``
        # and ``-->``, so its text starts directly with ``sol nom=...``.
        if child.tag is ET.Comment and (child.text or "").lstrip().startswith("sol "):
            root.remove(child)

    soil_updates = layer_updates = 0
    for source in sources:
        soil = deepcopy(soil_template)
        soil.set("nom", source.name)
        xml_parameters = _parameter_map(soil.iter(), f"soil {source.name!r}")
        missing = sorted(set(source.parameters) - set(xml_parameters))
        if missing:
            raise ValueError(
                f"soil {source.name!r}: parameters absent from XML template: "
                + ", ".join(missing)
            )
        for name, value in source.parameters.items():
            _set_xml_value(xml_parameters[name], value)
            soil_updates += 1

        tables = [child for child in list(soil) if child.tag == "tableau"]
        if len(tables) != 5:
            raise ValueError(
                f"{template_path}: template soil must contain 5 layers, found {len(tables)}"
            )
        for number, (table, values) in enumerate(zip(tables, source.layers), start=1):
            table.set("nom", f"layer {number}")
            columns = {column.get("nom"): column for column in table.findall("colonne")}
            missing_columns = sorted(set(values) - set(columns))
            if missing_columns:
                raise ValueError(
                    f"soil {source.name!r}, layer {number}: columns absent from XML "
                    f"template: {', '.join(missing_columns)}"
                )
            for name, value in values.items():
                columns[name].text = value
                layer_updates += 1
        root.append(soil)

    _write_xml(root, destination_path)
    return SoilTxtToXmlResult(destination_path, len(sources), soil_updates, layer_updates)


def _split_initialization_vector(path: Path, value: str, name: str) -> List[str]:
    values = value.split()
    if len(values) != 5:
        raise ValueError(f"{path}: {name} must contain 5 values, found {len(values)}")
    return values


def _read_initialization_txt(path: Path) -> Tuple[int, List[Dict[str, object]], Dict[str, List[str]], Dict[str, str]]:
    """Read the positional JavaSTICS initialization format used by STICS."""

    if not path.is_file():
        raise FileNotFoundError(path)
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    position = 0

    def take(description: str) -> str:
        nonlocal position
        if position >= len(lines):
            raise ValueError(f"{path}: missing {description}")
        value = lines[position].strip()
        position += 1
        return value

    def label(expected: str, aliases: Tuple[str, ...] = ()) -> None:
        actual = take(f"{expected} label").strip(":")
        if actual not in (expected,) + aliases:
            raise ValueError(f"{path}: expected {expected!r} label, found {actual!r}")

    label("nbplantes")
    try:
        plant_count = int(take("number of plants"))
    except ValueError as exc:
        raise ValueError(f"{path}: invalid number of plants") from exc
    if plant_count not in (1, 2):
        raise ValueError(f"{path}: expected 1 or 2 plants, found {plant_count}")

    plants: List[Dict[str, object]] = []
    scalar_names = (
        "stade0", "lai0", "magrain0", "zrac0", "code_acti_reserve",
        "maperenne0", "QNperenne0", "masecnp0", "QNplantenp0",
        "masec0", "QNplante0", "restemp0",
    )
    for _ in range(plant_count):
        label("plante")
        values = [take(name) for name in scalar_names[:4]]
        label("code_acti_reserve")
        values.extend(take(name) for name in scalar_names[4:])
        density_label = take("densinitial label").strip(":")
        if density_label != "densinitial":
            raise ValueError(f"{path}: expected 'densinitial' label, found {density_label!r}")
        densities = _split_initialization_vector(path, take("densinitial values"), "densinitial")
        plant = dict(zip(scalar_names, values))
        plant["densinitial"] = densities
        plants.append(plant)

    # The legacy one-plant file reserves a complete empty second plant block.
    if plant_count == 1:
        for _ in range(16):
            take("reserved second plant line")

    soil: Dict[str, List[str]] = {}
    for xml_name, text_name, aliases in (
        ("Hinitf", "Hinitf", ()),
        ("NO3initf", "NO3initf", ("NO3init",)),
        ("NH4initf", "NH4initf", ()),
    ):
        label(text_name, aliases)
        soil[xml_name] = _split_initialization_vector(path, take(f"{text_name} values"), text_name)

    snow: Dict[str, str] = {}
    if position < len(lines):
        label("snow")
        for name in ("Sdepth0", "Sdry0", "Swet0", "ps0"):
            label(name)
            snow[name] = take(f"{name} value")
    if any(line.strip() for line in lines[position:]):
        raise ValueError(f"{path}: unexpected data after initialization values")
    return plant_count, plants, soil, snow


def convert_initialization_txt_to_xml(
    source_txt: PathLike,
    template_xml: PathLike,
    destination_xml: PathLike,
) -> InitializationTxtToXmlResult:
    """Convert one positional STICS ``ficini.txt`` file to STICS V11 XML."""

    source_path, template_path = Path(source_txt), Path(template_xml)
    destination_path = Path(destination_xml)
    plant_count, source_plants, soil_values, snow_values = _read_initialization_txt(source_path)
    tree = _parse_template(template_path)
    root = tree.getroot()
    if root.tag != "initialisations":
        raise ValueError(
            f"{template_path}: expected an initialisations root, found {root.tag!r}"
        )
    count_element = root.find("nbplantes")
    if count_element is None:
        raise ValueError(f"{template_path}: missing nbplantes element")
    count_element.text = str(plant_count)
    xml_plants = root.findall("plante")
    if len(xml_plants) < plant_count:
        raise ValueError(
            f"{template_path}: {plant_count} plant blocks required, found {len(xml_plants)}"
        )

    plant_updates = 0
    for xml_plant, source in zip(xml_plants, source_plants):
        for name, value in source.items():
            if name == "densinitial":
                horizons = xml_plant.findall("densinitial/horizon")
                if len(horizons) != 5:
                    raise ValueError(f"{template_path}: densinitial must have 5 horizons")
                for horizon, horizon_value in zip(horizons, value):
                    horizon.text = str(horizon_value)
                    plant_updates += 1
            elif name == "code_acti_reserve":
                option = xml_plant.find("option[@nomParam='code_acti_reserve']")
                if option is None:
                    raise ValueError(f"{template_path}: missing code_acti_reserve option")
                option.set("choix", str(value))
                plant_updates += 1
            else:
                element = xml_plant.find(f".//{name}")
                if element is None:
                    raise ValueError(f"{template_path}: missing initialization value {name}")
                element.text = str(value)
                plant_updates += 1

    xml_soil = root.find("sol")
    if xml_soil is None:
        raise ValueError(f"{template_path}: missing sol initialization block")
    soil_updates = 0
    for name, values in soil_values.items():
        horizons = xml_soil.findall(f"{name}/horizon")
        if len(horizons) != 5:
            raise ValueError(f"{template_path}: {name} must have 5 horizons")
        for horizon, value in zip(horizons, values):
            horizon.text = value
            soil_updates += 1

    snow_updates = 0
    for name, value in snow_values.items():
        element = root.find(f"snow/{name}")
        if element is None:
            raise ValueError(f"{template_path}: missing snow initialization {name}")
        element.text = value
        snow_updates += 1

    _write_xml(root, destination_path)
    return InitializationTxtToXmlResult(
        destination_path, plant_count, plant_updates, soil_updates, snow_updates
    )


def _read_technical_txt(
    path: Path,
) -> Tuple[Dict[str, str], List[Tuple[List[str], List[List[str]]]]]:
    """Read the named-value fictec format and its ordered intervention tables."""

    if not path.is_file():
        raise FileNotFoundError(path)
    lines = [line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines()
             if line.strip()]
    parameters: Dict[str, str] = {}
    tables: List[Tuple[List[str], List[List[str]]]] = []
    position = 0
    while position < len(lines):
        name = lines[position]
        position += 1
        if position >= len(lines):
            raise ValueError(f"{path}: missing value after {name!r}")
        value = lines[position]
        position += 1
        if name == "nbinterventions":
            try:
                count = int(value)
            except ValueError as exc:
                raise ValueError(f"{path}: invalid nbinterventions value {value!r}") from exc
            if count < 0:
                raise ValueError(f"{path}: nbinterventions cannot be negative")
            header: List[str] = []
            rows: List[List[str]] = []
            if count:
                if position >= len(lines):
                    raise ValueError(f"{path}: missing intervention header")
                header = lines[position].split()
                position += 1
                for row_number in range(1, count + 1):
                    if position >= len(lines):
                        raise ValueError(f"{path}: missing intervention row {row_number}")
                    row = lines[position].split()
                    position += 1
                    if len(row) != len(header):
                        raise ValueError(
                            f"{path}: intervention row {row_number} contains {len(row)} "
                            f"values; expected {len(header)}"
                        )
                    rows.append(row)
            tables.append((header, rows))
        else:
            if name in parameters:
                raise ValueError(f"{path}: duplicate technical parameter {name!r}")
            parameters[name] = value
    return parameters, tables


def convert_technical_txt_to_xml(
    source_txt: PathLike,
    template_xml: PathLike,
    destination_xml: PathLike,
) -> TechnicalTxtToXmlResult:
    """Convert one STICS ``fictec`` management file to STICS V11 XML.

    Management types with zero interventions are represented by an empty
    ``ta`` table. Parameters absent from the TXT retain their template value,
    allowing files that do not expose every optional management formalism.
    """

    source_path, template_path = Path(source_txt), Path(template_xml)
    destination_path = Path(destination_xml)
    source_parameters, source_tables = _read_technical_txt(source_path)
    tree = _parse_template(template_path)
    root = tree.getroot()
    if root.tag != "fichiertec":
        raise ValueError(f"{template_path}: expected a fichiertec root, found {root.tag!r}")

    xml_parameters = _parameter_map(root.iter(), "technical template")
    unsupported = sorted(set(source_parameters) - set(xml_parameters))
    if unsupported:
        raise ValueError(
            "technical parameters from TXT absent from XML template: "
            + ", ".join(unsupported)
        )
    for name, value in source_parameters.items():
        _set_xml_value(xml_parameters[name], value)

    xml_tables = list(root.iter("ta"))
    if len(source_tables) != len(xml_tables):
        raise ValueError(
            f"{source_path}: found {len(source_tables)} intervention counts; "
            f"template requires {len(xml_tables)}"
        )
    intervention_count = 0
    for table_number, (table, (header, rows)) in enumerate(
        zip(xml_tables, source_tables), start=1
    ):
        expected_header_element = table.find("ta_entete")
        if expected_header_element is None:
            raise ValueError(f"{template_path}: table {table_number} has no ta_entete")
        expected_header = [column.get("nom") for column in expected_header_element.findall("colonne")]
        if rows and header != expected_header:
            raise ValueError(
                f"{source_path}: table {table_number} header {header!r} does not "
                f"match template header {expected_header!r}"
            )
        for intervention in list(table.findall("intervention")):
            table.remove(intervention)
        for row in rows:
            intervention = ET.Element("intervention", {"nb_colonnes": str(len(expected_header))})
            for name, value in zip(expected_header, row):
                column = ET.SubElement(intervention, "colonne", {"nom": str(name)})
                column.text = value
            table.append(intervention)
            intervention_count += 1
        table.set("nb_interventions", str(len(rows)))

    _write_xml(root, destination_path, standalone=False)
    return TechnicalTxtToXmlResult(
        destination_path, len(source_parameters), len(source_tables), intervention_count
    )


def _read_named_value_txt(path: Path, description: str) -> Dict[str, str]:
    if not path.is_file():
        raise FileNotFoundError(path)
    lines = [line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines()
             if line.strip()]
    if len(lines) % 2:
        raise ValueError(f"{path}: missing value after {lines[-1]!r}")
    values: Dict[str, str] = {}
    for position in range(0, len(lines), 2):
        name, value = lines[position], lines[position + 1]
        if name in values:
            raise ValueError(f"{path}: duplicate {description} parameter {name!r}")
        values[name] = value
    return values


def _read_named_value_occurrences(path: Path) -> List[Tuple[str, str]]:
    """Read name/value pairs without collapsing repeated parameter names."""

    if not path.is_file():
        raise FileNotFoundError(path)
    lines = [line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines()
             if line.strip()]
    if len(lines) % 2:
        raise ValueError(f"{path}: missing value after {lines[-1]!r}")
    return [(lines[index], lines[index + 1]) for index in range(0, len(lines), 2)]


def convert_general_parameters_txt_to_xml(
    source_txt: PathLike,
    template_xml: PathLike,
    destination_xml: PathLike,
    *,
    aliases: Optional[Dict[str, str]] = None,
) -> GeneralParametersTxtToXmlResult:
    """Convert ``tempopar*.sti`` to its general-parameters XML counterpart.

    Repeated names are matched occurrence by occurrence, in document order.
    This is required for indexed STICS collections such as soil textures,
    fertilizers, and the 21 residue types.
    """

    source_path, template_path = Path(source_txt), Path(template_xml)
    destination_path = Path(destination_xml)
    source_values = _read_named_value_occurrences(source_path)
    effective_aliases = {"humirac": "code_humirac"}
    if aliases:
        effective_aliases.update(aliases)

    tree = _parse_template(template_path)
    root = tree.getroot()
    if root.tag not in {"fichierpar", "fichierparamgen"}:
        raise ValueError(
            f"{template_path}: expected fichierpar or fichierparamgen root, "
            f"found {root.tag!r}"
        )
    xml_occurrences: Dict[str, List[ET.Element]] = {}
    for element in root.iter():
        name = element.get("nom") if element.tag == "param" else (
            element.get("nomParam") if element.tag == "option" else None
        )
        if name:
            xml_occurrences.setdefault(name, []).append(element)

    consumed: Dict[str, int] = {}
    for source_name, value in source_values:
        name = effective_aliases.get(source_name, source_name)
        index = consumed.get(name, 0)
        matches = xml_occurrences.get(name, [])
        if index >= len(matches):
            raise ValueError(
                f"{source_path}: occurrence {index + 1} of parameter {source_name!r} "
                f"has no counterpart in {template_path.name}"
            )
        _set_xml_value(matches[index], value)
        consumed[name] = index + 1

    missing = []
    for name, elements in xml_occurrences.items():
        used = consumed.get(name, 0)
        if used < len(elements):
            missing.append(f"{name} ({len(elements) - used} occurrence(s))")
    if missing:
        raise ValueError(
            f"{source_path}: XML occurrences without TXT values: " + ", ".join(missing)
        )

    _write_xml(root, destination_path, standalone=False)
    return GeneralParametersTxtToXmlResult(
        destination_path,
        len(source_values),
        len({effective_aliases.get(name, name) for name, _ in source_values}),
    )


def convert_station_txt_to_xml(
    source_txt: PathLike,
    template_xml: PathLike,
    destination_xml: PathLike,
    *,
    strict: bool = True,
) -> StationTxtToXmlResult:
    """Convert one alternating name/value STICS station file to V11 XML."""

    source_path, template_path = Path(source_txt), Path(template_xml)
    destination_path = Path(destination_xml)
    source_parameters = _read_named_value_txt(source_path, "station")
    tree = _parse_template(template_path)
    root = tree.getroot()
    if root.tag != "fichiersta":
        raise ValueError(f"{template_path}: expected a fichiersta root, found {root.tag!r}")
    xml_parameters = _parameter_map(root.iter(), "station template")
    defaults = _check_parameter_sets(
        source_parameters, xml_parameters, "station", strict
    )
    for name, value in source_parameters.items():
        _set_xml_value(xml_parameters[name], value)
    _write_xml(root, destination_path, standalone=False)
    return StationTxtToXmlResult(destination_path, len(source_parameters), defaults)


def generate_stics_observation_files(
    master_input_db: PathLike,
    destination_directory: PathLike,
    simulation_ids: Iterable[str],
    *,
    sowing_days: Dict[str, str],
) -> ObservationsResult:
    """Generate one semicolon-delimited STICS ``.obs`` file per simulation.

    STICS encodes multi-year ``maturitydate`` as ``365 * year_offset + DOY``
    (even across leap years). The emitted ``jul`` is the annual day of year,
    while ``imats`` preserves the cumulative ``maturitydate`` value.
    """

    database = Path(master_input_db)
    destination = Path(destination_directory)
    if not database.is_file():
        raise FileNotFoundError(database)
    destination.mkdir(parents=True, exist_ok=True)
    file_count = row_count = 0
    with sqlite3.connect(database) as connection:
        connection.row_factory = sqlite3.Row
        for simulation_id in simulation_ids:
            simulation = connection.execute(
                "SELECT StartYear FROM SimUnitList WHERE lower(idsim)=lower(?) LIMIT 1",
                (simulation_id,),
            ).fetchone()
            if simulation is None:
                raise ValueError(f"{database}: no SimUnitList row for {simulation_id}")
            observations = connection.execute(
                "SELECT maturitydate, maturitydate_calendar FROM dailyobs "
                "WHERE lower(idsim)=lower(?) ORDER BY maturitydate_calendar, maturitydate",
                (simulation_id,),
            ).fetchall()
            if not observations:
                continue
            if simulation_id not in sowing_days:
                raise ValueError(f"missing sowing day for observation file {simulation_id}")
            lines = ["ian;mo;jo;jul;iplts;imats"]
            start_year = int(simulation["StartYear"])
            for observation in observations:
                maturity = observation["maturitydate"]
                calendar = observation["maturitydate_calendar"]
                if calendar:
                    try:
                        maturity_date = date.fromisoformat(str(calendar))
                    except ValueError as exc:
                        raise ValueError(
                            f"{database}: invalid maturitydate_calendar {calendar!r} "
                            f"for {simulation_id}"
                        ) from exc
                    day_of_year = maturity_date.timetuple().tm_yday
                    expected_maturity = (
                        365 * (maturity_date.year - start_year) + day_of_year
                    )
                    if maturity is not None and int(maturity) != expected_maturity:
                        raise ValueError(
                            f"{database}: inconsistent maturity dates for {simulation_id}: "
                            f"maturitydate={maturity!r}, expected {expected_maturity} "
                            f"from calendar date {maturity_date}"
                        )
                elif maturity is not None:
                    zero_based = int(maturity) - 1
                    maturity_year = start_year + zero_based // 365
                    maturity_doy = zero_based % 365 + 1
                    maturity_date = date(maturity_year, 1, 1) + timedelta(
                        days=maturity_doy - 1
                    )
                else:
                    raise ValueError(f"{database}: missing maturity date for {simulation_id}")
                day_of_year = maturity_date.timetuple().tm_yday
                imats = int(maturity) if maturity is not None else (
                    365 * (maturity_date.year - start_year) + day_of_year
                )
                lines.append(
                    f"{maturity_date.year};{maturity_date.month};{maturity_date.day};"
                    f"{day_of_year};{sowing_days[simulation_id]};{imats}"
                )
                row_count += 1
            (destination / f"{simulation_id}.obs").write_text(
                "\n".join(lines) + "\n", encoding="utf-8"
            )
            file_count += 1
    return ObservationsResult(destination, file_count, row_count)


def _read_usm_txt(path: Path) -> Dict[str, str]:
    values = _read_named_value_txt(path, "USM")
    return {name.strip(":"): value for name, value in values.items()}


def _set_child_text(parent: ET.Element, name: str, value: str, context: str) -> None:
    element = parent.find(name)
    if element is None:
        raise ValueError(f"{context}: missing XML element {name!r}")
    element.text = str(value)


def _spreadsheet_value(value: Optional[str]):
    if value is None:
        return None
    text_value = str(value).strip()
    try:
        return int(text_value)
    except ValueError:
        try:
            return float(text_value)
        except ValueError:
            return text_value


def _xml_scalar_dict(root: ET.Element) -> Dict[str, object]:
    values: Dict[str, object] = {}
    for element in root.iter():
        if element.tag == "param" and element.get("nom"):
            values.setdefault(element.get("nom"), _spreadsheet_value(element.text))
        elif element.tag == "option" and element.get("nomParam"):
            values.setdefault(
                element.get("nomParam"), _spreadsheet_value(element.get("choix"))
            )
    return values


def generate_stics_summary_workbook(
    workspace_directory: PathLike,
    template_xlsx: PathLike,
    destination_xlsx: PathLike,
    *,
    general_parameters_xml: Optional[PathLike] = None,
    new_form_parameters_xml: Optional[PathLike] = None,
    master_input_db: Optional[PathLike] = None,
) -> Path:
    """Summarize a generated JavaSTICS workspace and all its dependencies."""

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to generate the STICS workbook") from exc

    workspace, template = Path(workspace_directory), Path(template_xlsx)
    destination = Path(destination_xlsx)
    if not (workspace / "usms.xml").is_file():
        raise FileNotFoundError(workspace / "usms.xml")
    if not template.is_file():
        raise FileNotFoundError(template)
    workbook = load_workbook(template)

    def reset_sheet(name: str, headers: Optional[List[str]] = None):
        sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)
        if sheet.max_row:
            sheet.delete_rows(1, sheet.max_row)
        if headers:
            sheet.append(headers)
        sheet.freeze_panes = "A2"
        if headers:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).coordinate}"
        return sheet

    original_headers = {
        name: [cell.value for cell in workbook[name][1] if cell.value is not None]
        for name in ("USMs", "Ini", "Soils", "Tec", "Station", "Obs")
    }
    sheets = {name: reset_sheet(name, headers) for name, headers in original_headers.items()}
    weather_headers = [
        "climate_file", "station", "year", "month", "day", "julian",
        "tmin", "tmax", "radiation", "etp", "rain", "wind", "vapour_pressure", "co2",
    ]
    sheets["Weather"] = reset_sheet("Weather", weather_headers)
    sheets["Plant"] = reset_sheet(
        "Plant", ["Plant_file", "scope", "cultivar", "parameter", "occurrence", "value"]
    )
    general_headers = [
        "source_file", "formalism", "type", "parameter", "occurrence", "value"
    ]
    sheets["GeneralParameters"] = reset_sheet("GeneralParameters", general_headers)
    sheets["NewFormParameters"] = reset_sheet("NewFormParameters", general_headers)

    if master_input_db is not None:
        database = Path(master_input_db)
        if not database.is_file():
            raise FileNotFoundError(database)
        summary_column_names = {
            "Planting": "iplts",
            "Emergence": "ilevs",
            "Ant": "iflos",
            "Mat": "imats",
            "Biom_ma": "masec(n)",
            "Yield": "mafruit",
            "GNumber": "chargefruit",
            "MaxLai": "laimax",
            "Nleac": "Qles",
            "SoilN": "QNapp",
            "CroN_ma": "QNplante",
            "CumE": "ces",
            "Transp": "cep",
        }
        summary_column_names_lower = {
            name.lower(): value for name, value in summary_column_names.items()
        }
        with sqlite3.connect(database) as connection:
            for table_name, sheet_name, renamed_columns in (
                ("SummaryOutput", "rapport", summary_column_names),
                ("SticsDailyOutput", "dailyOutput", {}),
            ):
                cursor = connection.execute(f'SELECT * FROM "{table_name}"')
                source_headers = [description[0] for description in cursor.description]
                headers = [
                    summary_column_names_lower.get(name.lower(), name)
                    if renamed_columns else name
                    for name in source_headers
                ]
                sheet = reset_sheet(sheet_name, headers)
                for row in cursor:
                    sheet.append(list(row))

            comparison_cursor = connection.execute(
                """
                SELECT s.IdSim, u.idPoint, d.maturitydate_calendar,
                       s.Mat, u.StartYear, s.SeasonOrder
                FROM SummaryOutput AS s
                INNER JOIN SimUnitList AS u
                    ON lower(u.idsim) = lower(s.IdSim)
                INNER JOIN dailyobs AS d
                    ON lower(d.idsim) = lower(s.IdSim)
                WHERE lower(s.Model) = 'stics'
                  AND d.maturitydate_calendar IS NOT NULL
                  AND s.Mat IS NOT NULL
                ORDER BY u.idPoint, d.maturitydate_calendar, s.IdSim
                """
            )
            comparison_sheet = reset_sheet(
                "MaturityComparison",
                [
                    "IdSim", "idPoint", "ObservedHarvestDate",
                    "SimulatedMaturityDate", "ErrorDays", "SeasonOrder",
                ],
            )
            site_ranges: Dict[str, List[int]] = {}
            for idsim, site, observed_text, imats, start_year, season_order in comparison_cursor:
                try:
                    observed_date = date.fromisoformat(str(observed_text))
                except ValueError as exc:
                    raise ValueError(
                        f"{database}: invalid observed harvest date {observed_text!r} "
                        f"for {idsim}"
                    ) from exc
                simulated_date = date(int(start_year), 1, 1) + timedelta(days=int(imats) - 1)
                comparison_sheet.append(
                    [
                        idsim, site, observed_date, simulated_date,
                        (simulated_date - observed_date).days, season_order,
                    ]
                )
                row_number = comparison_sheet.max_row
                site_ranges.setdefault(str(site), []).append(row_number)

            if site_ranges:
                from openpyxl.chart import Reference, ScatterChart, Series

                chart = ScatterChart()
                chart.title = "Simulated maturity versus observed harvest"
                chart.x_axis.title = "Observed harvest date"
                chart.y_axis.title = "Simulated maturity date"
                chart.x_axis.number_format = "yyyy-mm-dd"
                chart.y_axis.number_format = "yyyy-mm-dd"
                chart.height = 16
                chart.width = 26
                chart.legend.position = "r"
                for site_index, (site, rows) in enumerate(site_ranges.items()):
                    first_row, last_row = rows[0], rows[-1]
                    xvalues = Reference(
                        comparison_sheet, min_col=3, min_row=first_row, max_row=last_row
                    )
                    yvalues = Reference(
                        comparison_sheet, min_col=4, min_row=first_row, max_row=last_row
                    )
                    series = Series(yvalues, xvalues, title=site)
                    red, green, blue = colorsys.hsv_to_rgb(
                        site_index / len(site_ranges), 0.72, 0.82
                    )
                    colour = f"{round(red * 255):02X}{round(green * 255):02X}{round(blue * 255):02X}"
                    series.marker.symbol = "circle"
                    series.marker.size = 6
                    series.marker.graphicalProperties.solidFill = colour
                    series.marker.graphicalProperties.line.solidFill = colour
                    series.graphicalProperties.line.noFill = True
                    chart.series.append(series)
                comparison_sheet.add_chart(chart, "H2")

    usms_root = ET.parse(workspace / "usms.xml").getroot()
    usms = usms_root.findall("usm")
    usm_headers = original_headers["USMs"]
    for usm in usms:
        values = {child.tag: child.text for child in usm if isinstance(child.tag, str)}
        row = {name: values.get(name) for name in usm_headers}
        row["usm_name"] = usm.get("nom")
        for plant in usm.findall("plante"):
            suffix = plant.get("dominance")
            for field in ("fplt", "ftec", "flai", "fobs"):
                element = plant.find(field)
                row[f"{field}_{suffix}"] = element.text if element is not None else None
        sheets["USMs"].append([_spreadsheet_value(row.get(name)) for name in usm_headers])

    # Initialization files: one row per USM dependency.
    ini_headers = original_headers["Ini"]
    for filename in dict.fromkeys(usm.find("finit").text for usm in usms):
        root = ET.parse(workspace / filename).getroot()
        row: Dict[str, object] = {"Ini_name": filename, "nbplantes": int(root.find("nbplantes").text)}
        for plant_index, plant in enumerate(root.findall("plante"), start=1):
            suffix = f"_Crop{plant_index}"
            for child in plant:
                if child.tag == "option":
                    row[f"{child.get('nomParam')}{suffix}"] = _spreadsheet_value(child.get("choix"))
                    for parameter in child.iter():
                        if parameter is not child and isinstance(parameter.tag, str) and parameter.tag not in {"choix"}:
                            if parameter.tag not in {"option", "param"} and len(parameter) == 0:
                                row[f"{parameter.tag}{suffix}"] = _spreadsheet_value(parameter.text)
                elif child.tag == "densinitial":
                    for horizon in child.findall("horizon"):
                        row[f"densinitial_{horizon.get('nh')}{suffix}"] = _spreadsheet_value(horizon.text)
                elif len(child) == 0:
                    row[f"{child.tag}{suffix}"] = _spreadsheet_value(child.text)
        soil = root.find("sol")
        for vector in list(soil) if soil is not None else []:
            for horizon in vector.findall("horizon"):
                row[f"{vector.tag}_{horizon.get('nh')}"] = _spreadsheet_value(horizon.text)
        sheets["Ini"].append([row.get(name) for name in ini_headers])

    # Combined soils.
    soil_headers = original_headers["Soils"]
    soils_root = ET.parse(workspace / "sols.xml").getroot()
    for soil in soils_root.findall("sol"):
        row = {"Soil_name": soil.get("nom")}
        row.update(_xml_scalar_dict(soil))
        for layer_index, table in enumerate(soil.findall("tableau"), start=1):
            for column in table.findall("colonne"):
                row[f"{column.get('nom')}_{layer_index}"] = _spreadsheet_value(column.text)
        sheets["Soils"].append([row.get(name) for name in soil_headers])

    # Technical files and their intervention tables.
    tec_headers = original_headers["Tec"]
    tec_names = dict.fromkeys(
        plant.find("ftec").text for usm in usms for plant in usm.findall("plante")
        if plant.find("ftec") is not None and plant.find("ftec").text != "null"
    )
    for filename in tec_names:
        root = ET.parse(workspace / filename).getroot()
        row = {"Tec_name": filename}
        row.update(_xml_scalar_dict(root))
        tables = list(root.iter("ta"))
        if len(tables) >= 2:
            for index, intervention in enumerate(tables[1].findall("intervention"), start=1):
                for column in intervention.findall("colonne"):
                    row[f"{column.get('nom')}_{index}"] = _spreadsheet_value(column.text)
        if len(tables) >= 4:
            for index, intervention in enumerate(tables[3].findall("intervention"), start=1):
                columns = {c.get("nom"): c.text for c in intervention.findall("colonne")}
                row[f"julapN_{index}"] = _spreadsheet_value(columns.get("julapN_or_sum_upvt"))
                row[f"doseN_{index}"] = _spreadsheet_value(columns.get("absolute_value/%"))
                row[f"engrais_{index}"] = _spreadsheet_value(columns.get("engrais"))
        for table in tables[4:6]:
            for index, intervention in enumerate(table.findall("intervention"), start=1):
                for column in intervention.findall("colonne"):
                    row[f"{column.get('nom')}_{index}"] = _spreadsheet_value(column.text)
        sheets["Tec"].append([row.get(name) for name in tec_headers])

    station_headers = original_headers["Station"]
    station_names = dict.fromkeys(usm.find("fstation").text for usm in usms)
    for filename in station_names:
        root = ET.parse(workspace / filename).getroot()
        row = {"Sta_name": filename}
        row.update(_xml_scalar_dict(root))
        sheets["Station"].append([row.get(name) for name in station_headers])

    obs_headers = original_headers["Obs"]
    for usm in usms:
        filename = usm.find("plante[@dominance='1']/fobs").text
        if filename == "null" or not (workspace / filename).is_file():
            continue
        lines = (workspace / filename).read_text(encoding="utf-8-sig").splitlines()
        headers = lines[0].split(";")
        for line in lines[1:]:
            values = dict(zip(headers, line.split(";")))
            values["usm_name"] = usm.get("nom")
            sheets["Obs"].append([_spreadsheet_value(values.get(name)) for name in obs_headers])

    climate_names = dict.fromkeys(
        usm.find(field).text for usm in usms for field in ("fclim1", "fclim2")
    )
    for filename in climate_names:
        for line in (workspace / filename).read_text(encoding="utf-8-sig").splitlines():
            fields = line.split()
            if fields:
                sheets["Weather"].append(
                    [filename] + [_spreadsheet_value(value) for value in fields]
                )

    plant_names = dict.fromkeys(
        plant.find("fplt").text for usm in usms for plant in usm.findall("plante")
        if plant.find("fplt") is not None and plant.find("fplt").text != "null"
    )
    for filename in plant_names:
        root = ET.parse(workspace / "plant" / filename).getroot()
        cultivar_elements = {id(item) for variety in root.iter("variete") for item in variety.iter()}
        counts: Dict[Tuple[str, str, str], int] = {}
        for element in root.iter():
            name = element.get("nom") if element.tag == "param" else (
                element.get("nomParam") if element.tag == "option" else None
            )
            if not name:
                continue
            variety = next(
                (item for item in root.iter("variete") if id(element) in {id(x) for x in item.iter()}),
                None,
            ) if id(element) in cultivar_elements else None
            scope = "cultivar" if variety is not None else "species"
            cultivar = variety.get("nom") if variety is not None else None
            key = (scope, cultivar or "", name)
            counts[key] = counts.get(key, 0) + 1
            value = element.get("choix") if element.tag == "option" else element.text
            sheets["Plant"].append(
                [filename, scope, cultivar, name, counts[key], _spreadsheet_value(value)]
            )

    def append_general_parameters(sheet_name: str, xml_path: Optional[PathLike]) -> None:
        if xml_path is None:
            return
        path = Path(xml_path)
        if not path.is_file():
            raise FileNotFoundError(path)
        root = ET.parse(path).getroot()
        counts: Dict[str, int] = {}
        for formalism in root.findall(".//formalisme"):
            formalism_name = formalism.get("nom")
            for element in formalism.iter():
                name = element.get("nom") if element.tag == "param" else (
                    element.get("nomParam") if element.tag == "option" else None
                )
                if not name:
                    continue
                counts[name] = counts.get(name, 0) + 1
                value = element.get("choix") if element.tag == "option" else element.text
                sheets[sheet_name].append(
                    [
                        path.name,
                        formalism_name,
                        element.tag,
                        name,
                        counts[name],
                        _spreadsheet_value(value),
                    ]
                )

    append_general_parameters("GeneralParameters", general_parameters_xml)
    append_general_parameters("NewFormParameters", new_form_parameters_xml)

    for sheet in workbook.worksheets:
        if sheet.max_column:
            for column in sheet.columns:
                letter = column[0].column_letter
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 35)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def convert_usm_directory_to_xml(
    source_directory: PathLike,
    usms_template_xml: PathLike,
    destination_directory: PathLike,
    *,
    plant_template_xml: PathLike,
    initialization_template_xml: PathLike,
    soil_template_xml: PathLike,
    technical_template_xml: PathLike,
    station_template_xml: PathLike,
    master_input_db: Optional[PathLike] = None,
    summary_template_xlsx: Optional[PathLike] = None,
    general_parameters_xml: Optional[PathLike] = None,
    new_form_parameters_xml: Optional[PathLike] = None,
) -> UsmDirectoryToXmlResult:
    """Convert a directory containing one legacy STICS folder per USM.

    USM-specific XML filenames are prefixed by the unique USM directory name.
    Plant XML files are shared by all USMs using the same TXT ``codevar`` and
    written as ``plant/<codevar>_plt.xml``. Soils are combined in ``sols.xml``.
    Each ``climat.txt`` is split by year and written under the filename
    declared by ``fclim1``/``fclim2`` in ``new_travail.usm``.
    """

    source_root = Path(source_directory)
    destination = Path(destination_directory)
    if not source_root.is_dir():
        raise NotADirectoryError(source_root)
    usm_directories = sorted(
        path for path in source_root.iterdir()
        if path.is_dir() and (path / "new_travail.usm").is_file()
    )
    if not usm_directories:
        raise ValueError(f"{source_root}: no directory containing new_travail.usm")

    tree = _parse_template(Path(usms_template_xml))
    root = tree.getroot()
    if root.tag != "usms":
        raise ValueError(f"{usms_template_xml}: expected a usms root, found {root.tag!r}")
    template_usms = [child for child in list(root) if child.tag == "usm"]
    if not template_usms:
        raise ValueError(f"{usms_template_xml}: no usm block in template")
    usm_template = template_usms[0]
    for element in template_usms:
        root.remove(element)

    destination.mkdir(parents=True, exist_ok=True)
    plant_destination = destination / "plant"
    plant_destination.mkdir(parents=True, exist_ok=True)
    plant_filenames_by_cultivar: Dict[str, str] = {}
    written_climates: Dict[str, str] = {}
    soil_paths: List[Path] = []
    soil_names: List[str] = []
    associated_count = 0
    sowing_days: Dict[str, str] = {}
    observed_ids = set()
    if master_input_db is not None:
        database = Path(master_input_db)
        if not database.is_file():
            raise FileNotFoundError(database)
        with sqlite3.connect(database) as connection:
            observed_ids = {
                str(row[0]).lower()
                for row in connection.execute("SELECT DISTINCT idsim FROM dailyobs")
            }
    codesimul_values = {"culture": "0", "feuille": "1"}

    for usm_directory in usm_directories:
        usm_name = usm_directory.name
        values = _read_usm_txt(usm_directory / "new_travail.usm")
        required = {
            "codesimul", "nbplantes", "datedebut", "datefin", "finit", "nomsol",
            "fstation", "fclim1", "fclim2", "culturean", "fplt1", "ftec1", "flai1",
        }
        missing = sorted(required - set(values))
        if missing:
            raise ValueError(f"{usm_directory}: missing USM fields: {', '.join(missing)}")
        if values["codesimul"] not in codesimul_values:
            raise ValueError(f"{usm_directory}: unsupported codesimul {values['codesimul']!r}")
        try:
            plant_count = int(values["nbplantes"])
        except ValueError as exc:
            raise ValueError(f"{usm_directory}: invalid nbplantes") from exc
        if plant_count not in (1, 2):
            raise ValueError(f"{usm_directory}: expected 1 or 2 plants")

        filenames = {
            "finit": f"{usm_name}_ini.xml",
            "fstation": f"{usm_name}_sta.xml",
        }
        convert_initialization_txt_to_xml(
            usm_directory / values["finit"], initialization_template_xml,
            destination / filenames["finit"],
        )
        convert_station_txt_to_xml(
            usm_directory / values["fstation"], station_template_xml,
            destination / filenames["fstation"],
        )
        for plant_index in range(1, plant_count + 1):
            suffix = str(plant_index)
            for field in ("fplt", "ftec", "flai"):
                if f"{field}{suffix}" not in values:
                    raise ValueError(f"{usm_directory}: missing {field}{suffix}")
            source_plant = usm_directory / values[f"fplt{suffix}"]
            plant_data = SticsCultivarFile.read(source_plant)
            if len(plant_data.cultivar_names) != 1:
                raise ValueError(
                    f"{source_plant}: expected exactly one codevar for shared plant XML, "
                    f"found {len(plant_data.cultivar_names)}"
                )
            cultivar = plant_data.cultivar_names[0]
            if not cultivar.strip() or Path(cultivar).name != cultivar or any(
                separator in cultivar for separator in ("/", "\\")
            ):
                raise ValueError(f"{source_plant}: invalid codevar for filename: {cultivar!r}")
            plant_filename = f"{cultivar}_plt.xml"
            filenames[f"fplt{suffix}"] = plant_filename
            name_suffix = "" if plant_index == 1 else f"_{plant_index}"
            filenames[f"ftec{suffix}"] = f"{usm_name}{name_suffix}_tec.xml"
            if cultivar not in plant_filenames_by_cultivar:
                convert_plant_txt_to_xml(
                    source_plant, plant_template_xml,
                    plant_destination / plant_filename,
                )
                plant_filenames_by_cultivar[cultivar] = plant_filename
                associated_count += 1
            convert_technical_txt_to_xml(
                usm_directory / values[f"ftec{suffix}"], technical_template_xml,
                destination / filenames[f"ftec{suffix}"],
            )
            if plant_index == 1:
                technical_values, _ = _read_technical_txt(
                    usm_directory / values[f"ftec{suffix}"]
                )
                if "iplt0" not in technical_values:
                    raise ValueError(f"{usm_directory}: fictec has no iplt0")
                sowing_days[usm_name] = technical_values["iplt0"]
        associated_count += 2 + plant_count
        soil_paths.append(usm_directory / values["nomsol"])
        soil_names.append(usm_name)

        climate_by_year: Dict[int, List[str]] = {}
        climate_path = usm_directory / "climat.txt"
        for line_number, line in enumerate(
            climate_path.read_text(encoding="utf-8-sig").splitlines(), start=1
        ):
            if not line.strip():
                continue
            fields = line.split()
            if len(fields) < 2:
                raise ValueError(f"{climate_path}: invalid line {line_number}")
            try:
                year = int(fields[1])
            except ValueError as exc:
                raise ValueError(f"{climate_path}: invalid year on line {line_number}") from exc
            climate_by_year.setdefault(year, []).append(line)
        for climate_field in ("fclim1", "fclim2"):
            filename = values[climate_field]
            if Path(filename).name != filename:
                raise ValueError(f"{usm_directory}: climate filename must not contain a path")
            try:
                year = int(filename.rsplit(".", 1)[1])
            except (IndexError, ValueError) as exc:
                raise ValueError(f"{usm_directory}: climate file {filename!r} has no year suffix") from exc
            if year not in climate_by_year:
                raise ValueError(f"{climate_path}: no climate records for {year}")
            content = "\n".join(climate_by_year[year]) + "\n"
            previous = written_climates.get(filename)
            if previous is not None and previous != content:
                raise ValueError(f"conflicting climate contents for shared file {filename}")
            if previous is None:
                (destination / filename).write_text(content, encoding="utf-8")
                written_climates[filename] = content

        usm = deepcopy(usm_template)
        usm.set("nom", usm_name)
        for field in ("datedebut", "datefin", "culturean", "nbplantes"):
            _set_child_text(usm, field, values[field], usm_name)
        _set_child_text(usm, "codesimul", codesimul_values[values["codesimul"]], usm_name)
        _set_child_text(usm, "finit", filenames["finit"], usm_name)
        _set_child_text(usm, "nomsol", usm_name, usm_name)
        _set_child_text(usm, "fstation", filenames["fstation"], usm_name)
        _set_child_text(usm, "fclim1", values["fclim1"], usm_name)
        _set_child_text(usm, "fclim2", values["fclim2"], usm_name)
        xml_plants = usm.findall("plante")
        if len(xml_plants) < 2:
            raise ValueError(f"{usms_template_xml}: USM template requires two plante blocks")
        for index, plant in enumerate(xml_plants, start=1):
            if index <= plant_count:
                suffix = str(index)
                _set_child_text(plant, "fplt", filenames[f"fplt{suffix}"], usm_name)
                _set_child_text(plant, "ftec", filenames[f"ftec{suffix}"], usm_name)
                _set_child_text(plant, "flai", values[f"flai{suffix}"], usm_name)
            else:
                for field in ("fplt", "ftec", "flai"):
                    _set_child_text(plant, field, "null", usm_name)
            observation_name = (
                f"{usm_name}.obs"
                if index == 1 and usm_name.lower() in observed_ids
                else "null"
            )
            _set_child_text(plant, "fobs", observation_name, usm_name)
        root.append(usm)

    convert_soil_txt_to_xml(
        soil_paths, soil_template_xml, destination / "sols.xml", soil_names=soil_names
    )
    observation_result = ObservationsResult(destination, 0, 0)
    if master_input_db is not None:
        observation_result = generate_stics_observation_files(
            master_input_db,
            destination,
            [name for name in soil_names if name.lower() in observed_ids],
            sowing_days=sowing_days,
        )
    _write_xml(root, destination / "usms.xml", standalone=False)
    summary_path = None
    if summary_template_xlsx is not None:
        summary_path = generate_stics_summary_workbook(
            destination,
            summary_template_xlsx,
            destination / Path(summary_template_xlsx).name,
            general_parameters_xml=general_parameters_xml,
            new_form_parameters_xml=new_form_parameters_xml,
            master_input_db=master_input_db,
        )
    return UsmDirectoryToXmlResult(
        destination, len(usm_directories), len(soil_paths), len(written_climates),
        associated_count + 1,
        observation_result.files,
        summary_path,
    )


def convert_stics_txt_to_xml(
    source_txt: Union[PathLike, Iterable[PathLike]],
    template_xml: PathLike,
    destination_xml: PathLike,
    *,
    file_type: str = "plant",
    template_cultivar: Optional[str] = None,
    strict: bool = True,
    template_soil: Optional[str] = None,
) -> Union[
    TxtToXmlResult,
    SoilTxtToXmlResult,
    InitializationTxtToXmlResult,
    TechnicalTxtToXmlResult,
    StationTxtToXmlResult,
    GeneralParametersTxtToXmlResult,
]:
    """Dispatch a STICS TXT-to-XML conversion by input file type."""

    if file_type == "soil":
        sources = [source_txt] if isinstance(source_txt, (str, Path)) else source_txt
        return convert_soil_txt_to_xml(
            sources, template_xml, destination_xml, template_soil=template_soil
        )
    if file_type == "initialization":
        if not isinstance(source_txt, (str, Path)):
            raise TypeError("initialization conversion accepts exactly one source file")
        return convert_initialization_txt_to_xml(source_txt, template_xml, destination_xml)
    if file_type == "technical":
        if not isinstance(source_txt, (str, Path)):
            raise TypeError("technical conversion accepts exactly one source file")
        return convert_technical_txt_to_xml(source_txt, template_xml, destination_xml)
    if file_type == "station":
        if not isinstance(source_txt, (str, Path)):
            raise TypeError("station conversion accepts exactly one source file")
        return convert_station_txt_to_xml(
            source_txt, template_xml, destination_xml, strict=strict
        )
    if file_type == "general_parameters":
        if not isinstance(source_txt, (str, Path)):
            raise TypeError("general-parameters conversion accepts exactly one source file")
        return convert_general_parameters_txt_to_xml(
            source_txt, template_xml, destination_xml
        )
    if file_type != "plant":
        raise ValueError(f"unsupported STICS TXT file type: {file_type!r}")
    if not isinstance(source_txt, (str, Path)):
        raise TypeError("plant conversion accepts exactly one source file")
    return convert_plant_txt_to_xml(
        source_txt,
        template_xml,
        destination_xml,
        template_cultivar=template_cultivar,
        strict=strict,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", nargs="+", type=Path, help="STICS TXT file(s)")
    parser.add_argument("--template", required=True, type=Path, help="matching STICS XML template")
    parser.add_argument("--output", required=True, type=Path, help="destination XML file")
    parser.add_argument("--template-cultivar", help="variete block to use as structural template")
    parser.add_argument("--template-soil", help="sol block to use as structural template")
    parser.add_argument(
        "--file-type",
        choices=(
            "plant", "soil", "initialization", "technical", "station",
            "general_parameters",
        ),
        default="plant",
    )
    parser.add_argument(
        "--allow-template-defaults",
        action="store_true",
        help="retain XML defaults for parameters absent from TXT",
    )
    args = parser.parse_args()

    if args.file_type == "plant":
        if len(args.source) != 1:
            parser.error("plant conversion requires exactly one source file")
        result = convert_plant_txt_to_xml(
            args.source[0], args.template, args.output,
            template_cultivar=args.template_cultivar,
            strict=not args.allow_template_defaults,
        )
        print(f"{args.source[0].name}: {result.cultivars} cultivar(s) -> {result.destination}")
    elif args.file_type == "soil":
        result = convert_soil_txt_to_xml(
            args.source, args.template, args.output, template_soil=args.template_soil
        )
        print(f"{result.soils} soil(s) -> {result.destination}")
    elif args.file_type == "initialization":
        if len(args.source) != 1:
            parser.error("initialization conversion requires exactly one source file")
        result = convert_initialization_txt_to_xml(args.source[0], args.template, args.output)
        print(f"{result.plants} plant(s) initialized -> {result.destination}")
    elif args.file_type == "technical":
        if len(args.source) != 1:
            parser.error("technical conversion requires exactly one source file")
        result = convert_technical_txt_to_xml(args.source[0], args.template, args.output)
        print(
            f"{result.parameters_updated} parameter(s), {result.interventions} "
            f"intervention(s) -> {result.destination}"
        )
    elif args.file_type == "station":
        if len(args.source) != 1:
            parser.error("station conversion requires exactly one source file")
        result = convert_station_txt_to_xml(
            args.source[0], args.template, args.output,
            strict=not args.allow_template_defaults,
        )
        print(f"{result.parameters_updated} station parameter(s) -> {result.destination}")
    else:
        if len(args.source) != 1:
            parser.error("general-parameters conversion requires exactly one source file")
        result = convert_general_parameters_txt_to_xml(
            args.source[0], args.template, args.output
        )
        print(f"{result.values_updated} general parameter value(s) -> {result.destination}")


if __name__ == "__main__":
    main()
